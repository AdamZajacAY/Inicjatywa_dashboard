#!/usr/bin/env python3
"""
Jednorazowy migrator Baza_Projektow.xlsx -> baza_projektow.db (SQLite).

Po co to jest: dashboard przechodzi z architektury "Excel jako baza danych" na
prawdziwy backend (server.py) trzymajacy dane w SQLite. Ten skrypt czyta
dzisiejszy plik Excela (ta sama struktura co generuje generuj_baze.py) i
zaladowuje go do nowej bazy SQLite wedlug schema.sql.

Idempotentny: kazde uruchomienie czysci tabele i wczytuje na nowo z Excela -
bezpieczny do wielokrotnego odpalenia w trakcie przejscia na nowa architekture.

Uruchomienie:  python3 excel_to_sqlite.py
Wejscie:       Baza_Projektow.xlsx (w tym samym katalogu)
Wyjscie:       baza_projektow.db (w tym samym katalogu)
"""

import datetime
import os
import sqlite3
import sys

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(ROOT, "Baza_Projektow.xlsx")
DB_PATH = os.path.join(ROOT, "baza_projektow.db")
SCHEMA_PATH = os.path.join(ROOT, "schema.sql")

# (nazwa arkusza Excel, nazwa tabeli SQL, lista kolumn w kolejnosci naglowka Excela)
SHEETS = [
    ("Projekty", "projekty", [
        "ID_Projektu", "Nazwa", "Typ_projektu", "Funkcja_biura", "Segment", "Owner", "Kierownik_projektu",
        "Status", "Faza", "Priorytet", "RAG_Status", "Tagi", "Data_rozpoczecia", "Data_zakonczenia_planowana",
        "Data_zakonczenia_rzeczywista", "Procent_postepu", "Budzet_calkowity", "Budzet_wydany", "Waluta",
        "Przychod_planowany", "Przychod_rzeczywisty", "Szacowane_roboczogodziny", "Stawka_godzinowa_srednia",
        "Lokalizacja_Adres", "Miasto", "Powierzchnia_m2", "Liczba_jednostek", "Inwestor_Klient", "Opis",
        "Link_do_dokumentacji", "Data_ostatniej_aktualizacji", "Komentarz",
    ]),
    ("Zespol", "zespol", [
        "ID_Osoby", "Imie_i_nazwisko", "Stanowisko_Rola", "Dzial", "Email", "Telefon",
        "Dostepnosc_FTE_procent", "Stawka_godzinowa", "Data_dolaczenia", "Aktywny",
    ]),
    ("Przypisania", "przypisania", [
        "ID_Przypisania", "ID_Projektu", "ID_Osoby", "Rola_w_projekcie", "Procent_zaangazowania",
        "Data_od", "Data_do", "Status",
    ]),
    ("Harmonogram", "harmonogram", [
        "ID_Zadania", "ID_Projektu", "Nazwa_zadania", "Kategoria", "ID_Osoby_odpowiedzialnej",
        "Data_start_plan", "Data_koniec_plan", "Data_start_rzeczywista", "Data_koniec_rzeczywista",
        "Procent_ukonczenia", "ID_Zadania_poprzedzajacego", "Kamien_milowy", "Status", "Priorytet", "Uwagi",
    ]),
    ("Zadania_Tickety", "zadania_tickety", [
        "ID_Tickietu", "ID_Projektu", "ID_Etapu", "Tytul", "Opis", "ID_Osoby_przypisanej",
        "ID_Podwykonawcy", "Wycena_podwykonawcy",
        "Data_utworzenia", "Termin", "Szacowane_roboczogodziny", "Rzeczywiste_roboczogodziny",
        "Priorytet", "Status", "Data_zakonczenia",
    ]),
    ("Kamienie_milowe", "kamienie_milowe", [
        "ID_Kamienia", "ID_Projektu", "Nazwa_kamienia", "Data_planowana", "Data_rzeczywista",
        "Status", "ID_Osoby_odpowiedzialnej",
    ]),
    ("Podwykonawcy", "podwykonawcy", [
        "ID_Podwykonawcy", "Nazwa", "Branza", "Typ_wspolpracy", "Osoba_kontaktowa", "Email",
        "Telefon", "NIP", "Miasto", "Ocena", "Status", "Uwagi",
    ]),
    ("Przypisania_Podwykonawcow", "przypisania_podwykonawcow", [
        "ID_Przypisania_Podw", "ID_Projektu", "ID_Podwykonawcy", "Branza", "Zakres_prac",
        "Data_od", "Data_do", "Wartosc_umowy", "Waluta", "Status", "Uwagi",
    ]),
    ("Ryzyka_i_Problemy", "ryzyka_i_problemy", [
        "ID", "ID_Projektu", "Typ", "Opis", "Kategoria", "Prawdopodobienstwo", "Wplyw",
        "Priorytet", "ID_Osoby_wlasciciela", "Plan_mitygacji", "Status", "Data_identyfikacji", "Data_zamkniecia",
    ]),
    ("Raporty_statusowe", "raporty_statusowe", [
        "Data_raportu", "ID_Projektu", "RAG_Status", "Procent_postepu",
        "Budzet_wydany_skumulowany", "Kluczowe_osiagniecia", "Kluczowe_problemy", "Nastepne_kroki", "Autor_raportu",
    ]),
]


def cell_value(v):
    """Normalizuje wartosc komorki Excela do typu zapisywalnego w SQLite."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return v


def read_sheet_rows(ws, headers):
    header_row = [c.value for c in ws[1]]
    col_index = {name: header_row.index(name) for name in headers if name in header_row}
    missing = [name for name in headers if name not in header_row]
    if missing:
        print(f"  UWAGA: arkusz '{ws.title}' nie ma kolumn: {missing}")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append([cell_value(row[col_index[name]]) if name in col_index else None for name in headers])
    return rows


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"Nie znaleziono {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    with open(SCHEMA_PATH, encoding="utf-8") as f:
        conn.executescript(f.read())

    print("=" * 60)
    for sheet_name, table_name, headers in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"{sheet_name}: arkusz nie istnieje w pliku - pomijam")
            continue
        ws = wb[sheet_name]
        rows = read_sheet_rows(ws, headers)
        placeholders = ", ".join("?" for _ in headers)
        col_list = ", ".join(headers)
        conn.executemany(
            f"INSERT INTO {table_name} ({col_list}) VALUES ({placeholders})",
            rows,
        )
        print(f"{table_name:<30} {len(rows):>4} wierszy  (z arkusza {sheet_name})")
    conn.commit()

    print("=" * 60)
    print("Weryfikacja (COUNT w SQLite):")
    for _, table_name, _ in SHEETS:
        n = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
        print(f"  {table_name:<30} {n}")
    conn.close()
    print("=" * 60)
    print(f"Zapisano: {DB_PATH}")


if __name__ == "__main__":
    main()
