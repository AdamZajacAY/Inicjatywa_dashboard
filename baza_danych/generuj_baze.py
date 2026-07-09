#!/usr/bin/env python3
"""
Generuje plik Baza_Projektow.xlsx - baze danych portfela projektow biura architektonicznego
Inicjatywa Projektowa (projekty koncepcyjne, budowlane, wykonawcze, nadzory autorskie, konkursy,
analizy urbanistyczne).

Uruchomienie:  python3 generuj_baze.py
Wynik:         Baza_Projektow.xlsx (w tym samym katalogu)

Struktura pliku jest opisana w ../analiza/Analiza_systemu_PMO.md
Dashboard (../dashboard/) czyta WYLACZNIE surowe dane (ID-based), nie formuly,
dzieki czemu dziala niezaleznie od tego czy plik byl otwierany w Excelu.

UWAGA: Projekty/Zespol/Przypisania w tym pliku to REALNE dane zebrane z notatek
zespolu (lipiec 2026) - nie przykladowe dane demonstracyjne. Harmonogram, Kamienie_milowe,
Podwykonawcy, Przypisania_Podwykonawcow, Ryzyka_i_Problemy i Raporty_statusowe sa celowo
puste (sama struktura + walidacja) - do uzupelnienia na biezaco przez zespol w dashboardzie.
"""

import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# ---------------------------------------------------------------------------
# Style pomocnicze
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FONT_GREEN = Font(color="006100")
FONT_YELLOW = Font(color="9C6500")
FONT_RED = Font(color="9C0006")

wb = Workbook()
wb.remove(wb.active)


def add_sheet(name):
    return wb.create_sheet(name)


def write_header(ws, headers, row=1):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.row_dimensions[row].height = 32


def write_rows(ws, rows, start_row=2):
    for r_off, row in enumerate(rows):
        for c_off, val in enumerate(row, start=1):
            cell = ws.cell(row=start_row + r_off, column=c_off, value=val)
            cell.border = BORDER
    return start_row + len(rows) - 1  # ostatni wypelniony wiersz (moze byc < start_row jesli brak danych)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_table(ws, name, headers, first_row, last_row):
    if last_row <= first_row:
        return  # brak danych - nie twórz obiektu Table (Excel wymaga >=1 wiersza danych)
    ref = f"A{first_row}:{get_column_letter(len(headers))}{last_row}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
    ws.add_table(tbl)


def dropdown(ws, col_letter, first_row, last_row, source_range, sheet="Slowniki"):
    dv = DataValidation(type="list", formula1=f"={sheet}!{source_range}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def safe_last(n, minimum=2):
    """Gwarantuje, ze zakres A2:A{n} nigdy nie jest odwrocony, nawet gdy arkusz jest pusty."""
    return max(n, minimum)


# ===========================================================================
# 1) SLOWNIKI  (listy do walidacji danych - jedno miejsce prawdy)
# ===========================================================================
ws_dict = add_sheet("Slowniki")
dict_cols = {
    "Typy_projektow": ["Projekt koncepcyjny", "Analiza urbanistyczna", "Projekt budowlany",
                        "Projekt wykonawczy", "Nadzor autorski", "Konkurs", "Projekt techniczny (PT)", "Inne"],
    "Funkcje_biura": ["Projektant wiodacy", "Nadzor autorski", "Analiza/doradztwo", "Uczestnik konkursu", "Koordynacja branzowa"],
    "Statusy_projektu": ["Planowanie", "W realizacji", "Wstrzymany", "Zakonczony", "Anulowany"],
    "Fazy": ["Koncepcja", "Analiza", "Projektowanie", "Pozwolenia/Przetarg", "Budowa", "Zakonczenie",
             "Konkurs - etap studialny", "Konkurs - etap II"],
    "RAG": ["Zielony", "Zolty", "Czerwony"],
    "Priorytety": ["Wysoki", "Sredni", "Niski"],
    "Role_w_projekcie": ["Sponsor", "Owner", "Kierownik projektu", "Czlonek zespolu", "Wsparcie/Konsultant"],
    "Dzialy": ["Architekci", "Specjalisci", "Kierownictwo projektow", "PMO", "Prawny", "Finansowy", "Marketing/Sprzedaz", "Zarzad"],
    "Branze_podwykonawcow": ["Elektryczna", "Sanitarna/Hydrauliczna", "Gazowa", "Wentylacja i klimatyzacja",
                              "Konstrukcyjna", "Drogowa/Infrastruktura", "Teletechniczna/IT", "Przeciwpozarowa", "Inna"],
    "Typy_wspolpracy_podwykonawcow": ["Projektant branzowy", "Wykonawca robot", "Dostawca", "Konsultant"],
    "Oceny_podwykonawcow": ["Wysoka", "Srednia", "Niska", "Brak oceny"],
    "Statusy_podwykonawcow": ["Aktywny", "Nieaktywny", "Zweryfikowany", "Czarna lista"],
    "Statusy_przypisania_podwykonawcy": ["Planowany", "Aktywny", "Zakonczony", "Wstrzymany"],
    "Kategorie_ryzyk": ["Prawne", "Finansowe", "Techniczne", "Harmonogramowe", "Zasoby", "Srodowiskowe",
                         "Proceduralne/Przetargowe"],
    "Typ_ryzyka": ["Ryzyko", "Problem"],
    "Status_zadania": ["Nie rozpoczete", "W trakcie", "Zakonczone", "Opoznione"],
    "Status_ryzyka": ["Otwarte", "W trakcie", "Zamkniete"],
    "Segmenty": ["Mieszkaniowy", "Komercyjny", "Publiczny", "Zielen"],
    "Statusy_tickietow": ["Backlog", "W tym tygodniu", "W trakcie", "Do przegladu", "Zrobione", "Zablokowane", "Zarchiwizowane"],
    "Kategorie_zadan": ["Koncepcja", "Konsultacje", "Projektowanie", "Rysunki wykonawcze",
                         "Dokumentacja przetargowa", "Pozwolenia/Uzgodnienia", "Nadzor autorski",
                         "Koordynacja branzowa", "Wizja lokalna/Spotkanie", "Prezentacja", "Administracja/Inne"],
}
headers = list(dict_cols.keys())
write_header(ws_dict, headers)
maxlen = max(len(v) for v in dict_cols.values())
for r in range(maxlen):
    for c, key in enumerate(headers, start=1):
        vals = dict_cols[key]
        if r < len(vals):
            ws_dict.cell(row=r + 2, column=c, value=vals[r]).border = BORDER
autosize(ws_dict, [24] * len(headers))
ws_dict.sheet_properties.tabColor = "808080"


def dict_range(name, n=None):
    col_idx = headers.index(name) + 1
    col_letter = get_column_letter(col_idx)
    n = n or len(dict_cols[name])
    return f"${col_letter}$2:${col_letter}${n + 1}"


# ===========================================================================
# 2) ZESPOL (osoby zebrane z notatek 4 lideröw zespolow: Jan B, Wojtek F, Grzegorz K, Monika Ch)
# ===========================================================================
ws_team = add_sheet("Zespol")
team_headers = ["ID_Osoby", "Imie_i_nazwisko", "Stanowisko_Rola", "Dzial", "Email", "Telefon",
                "Dostepnosc_FTE_procent", "Stawka_godzinowa", "Data_dolaczenia", "Aktywny"]
write_header(ws_team, team_headers)
team_rows = [
    ["P01", "Jan B", "Architekt / Lider zespolu", "Kierownictwo projektow", None, None, 100, None, "Tak"],
    ["P02", "Toan C", "Architekt", "Architekci", None, None, 100, None, "Tak"],
    ["P03", "Weronika S", "Architekt", "Architekci", None, None, 100, None, "Tak"],
    ["P04", "Wojtek F", "Architekt / Lider zespolu", "Kierownictwo projektow", None, None, 100, None, "Tak"],
    ["P05", "Marcin", "Architekt (wsparcie)", "Architekci", None, None, 100, None, "Tak"],
    ["P06", "Adrian", "Architekt", "Architekci", None, None, 100, None, "Tak"],
    ["P07", "Gosia", "Architekt", "Architekci", None, None, 100, None, "Tak"],
    ["P08", "Grzegorz K", "Architekt / Lider zespolu (Nadzor autorski)", "Kierownictwo projektow", None, None, 100, None, "Tak"],
    ["P09", "Malgorzata LB", "Specjalista (wsparcie)", "Specjalisci", None, None, 100, None, "Tak"],
    ["P10", "Monika Ch", "Architekt / Lider zespolu", "Kierownictwo projektow", None, None, 100, None, "Tak"],
    ["P11", "Mikolaj Z", "Specjalista (kontakt z inwestorem)", "Specjalisci", None, None, 100, None, "Tak"],
    ["P12", "Beata Zdral", "Specjalista (wsparcie, dochodzaco)", "Specjalisci", None, None, 100, None, "Tak"],
]
# Wstrzykniecie wartosci Stawka_godzinowa (None) w istniejace wiersze - nagłowek juz jest
# poprawnie zdefiniowany w team_headers powyzej (przed write_header). Puste, bo realne stawki
# nie byly czescia notatek zrodlowych - do uzupelnienia w dashboardzie.
_STAWKA_IDX = team_headers.index("Stawka_godzinowa")
for _row in team_rows:
    _row.insert(_STAWKA_IDX, None)
assert all(len(_row) == len(team_headers) for _row in team_rows), "Zespol: niezgodna liczba kolumn po wstrzyknieciu Stawka_godzinowa"

last_team = write_rows(ws_team, team_rows)
add_table(ws_team, "TabZespol", team_headers, 1, last_team)
autosize(ws_team, [10, 20, 34, 26, 24, 14, 14, 14, 16, 10])
dropdown(ws_team, "D", 2, 200, dict_range("Dzialy"))
for r in range(2, last_team + 1):
    ws_team.cell(row=r, column=_STAWKA_IDX + 1).number_format = "#,##0.00"
ws_team.sheet_properties.tabColor = "1F3864"

# ===========================================================================
# 3) PROJEKTY (21 realnych projektow - stan na 2026-07-07)
# ===========================================================================
ws_proj = add_sheet("Projekty")
proj_headers = [
    "ID_Projektu", "Nazwa", "Typ_projektu", "Funkcja_biura", "Segment", "Owner", "Kierownik_projektu",
    "Status", "Faza", "Priorytet", "RAG_Status", "Tagi",
    "Data_rozpoczecia", "Data_zakonczenia_planowana", "Data_zakonczenia_rzeczywista",
    "Procent_postepu", "Budzet_calkowity", "Budzet_wydany", "Waluta",
    "Przychod_planowany", "Przychod_rzeczywisty",
    "Szacowane_roboczogodziny", "Stawka_godzinowa_srednia",
    "Lokalizacja_Adres", "Miasto", "Powierzchnia_m2", "Liczba_jednostek",
    "Inwestor_Klient", "Opis", "Link_do_dokumentacji", "Data_ostatniej_aktualizacji", "Komentarz",
]
write_header(ws_proj, proj_headers)
U = dt.date(2026, 7, 7)  # data kompilacji danych ("Data_ostatniej_aktualizacji")

proj_rows = [
    ["JB-01", "Slowackiego 108", "Projekt wykonawczy", "Projektant wiodacy", None, "Jan B", "Jan B",
     "W realizacji", "Projektowanie", None, "Zielony",
     None, None, None, None, None, None, "PLN", None, None,
     "ul. Slowackiego 108", None, None, None,
     "GH development", "Projekt budowlany i projekt wykonawczy budynku dla GH development.",
     None, U, "[Zrodlo: Jan B] Zespol: Jan B (lider), Toan C, Weronika S."],

    ["JB-02", "Zabi Kruk 14", "Nadzor autorski", "Nadzor autorski", None, "Jan B", "Jan B",
     "W realizacji", "Budowa", None, "Zielony",
     None, None, None, None, None, None, "PLN", None, None,
     "ul. Zabi Kruk 14", None, None, None,
     "GH development", "Nadzor autorski na budowie dla GH development.",
     None, U, "[Zrodlo: Jan B] Zespol: Jan B (lider), Toan C."],

    ["JB-03", "Dom Pomocy Spolecznej - Maslow k. Kielc", "Projekt koncepcyjny", "Projektant wiodacy", "Publiczny", "Jan B", "Jan B",
     "Wstrzymany", "Koncepcja", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Maslow", None, None,
     "Inwestor prywatny", "Projekt koncepcyjny domu pomocy spolecznej.",
     None, U, "[Zrodlo: Jan B] Wstrzymany (on hold), przyczyna nie podana w notatce zrodlowej."],

    ["JB-04", "Zabudowa wielorodzinna - Osiedle Siekierki", "Analiza urbanistyczna", "Analiza/doradztwo", "Mieszkaniowy", None, None,
     "W realizacji", "Analiza", None, "Zielony",
     None, None, None, None, None, None, "PLN", None, None,
     "Osiedle Siekierki", "Warszawa", None, None,
     None, "Analiza urbanistyczna pod zabudowe wielorodzinna.",
     None, U, "[Zrodlo: Jan B] W notatce zrodlowej osoba oznaczona jako 'J' - prawdopodobnie Jan B, wymaga potwierdzenia. Inwestor nie podany - do uzupelnienia."],

    ["JB-05", "TBS Kolska 10", "Nadzor autorski", "Nadzor autorski", "Mieszkaniowy", "Jan B", "Jan B",
     "Wstrzymany", "Budowa", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     "ul. Kolska 10", "Warszawa", None, None,
     "TBS Warszawa Polnoc", "Nadzor autorski na budowie dla TBS Warszawa Polnoc.",
     None, U, "[Zrodlo: Jan B] Zespol: Jan B (lider), Toan C. Wstrzymany (on hold)."],

    ["WF-01", "Wieliszew", "Nadzor autorski", "Nadzor autorski", None, "Wojtek F", "Wojtek F",
     "W realizacji", "Budowa", None, "Zielony",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Wieliszew", None, None,
     None, "Nadzor autorski na budowie.",
     None, U, "[Zrodlo: Wojtek F] Przeglad kart materialowych zaproponowanych przez budowe co 1-2 tyg., wprowadzanie korekt / uzgadnianie zmian z inwestorem lub budowa. Same podroze na budowe odbywaja sie juz bardzo rzadko."],

    ["WF-02", "IChP", "Inne", None, None, "Wojtek F", "Wojtek F",
     "W realizacji", "Zakonczenie", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     None, None, None, None,
     None, "Temat konczacy sie - dokladny typ i zakres projektu niejasny w notatce zrodlowej ('trudno powiedziec').",
     None, U, "[Zrodlo: Wojtek F] Zespol: Wojtek F (lider), Marcin (do pomocy). Temat konczacy sie, ale ma swoje mankamenty i jeszcze wracaja niezakonczone rzeczy."],

    ["WF-03", "GHS_3L - szklarnia", "Projekt budowlany", "Projektant wiodacy", "Komercyjny", "Wojtek F", "Wojtek F",
     "Wstrzymany", "Projektowanie", None, "Zolty",
     None, dt.date(2026, 7, 15), None, None, None, None, "PLN", None, None,
     None, None, None, None,
     None, "Projekt budowlany szklarni.",
     None, U, "[Zrodlo: Wojtek F] Zespol: Wojtek F (lider), Marcin. Na ten moment wstrzymany projekt do okolo 15.07.2026."],

    ["WF-04", "SKA", "Projekt wykonawczy", "Projektant wiodacy", None, "Wojtek F", "Wojtek F",
     "W realizacji", "Projektowanie", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     None, None, None, None,
     None, "Projekt budowlany i projekt wykonawczy.",
     None, U, "[Zrodlo: Wojtek F] Zespol: Wojtek F (lider), Adrian, Marcin (dorywczo), + osoba do wnetrz (jeszcze do znalezienia). Dosc trudny projekt, ktory ma sporo mankamentow i sporo rzeczy do zrobienia."],

    ["WF-05", "NIK", "Projekt budowlany", "Projektant wiodacy", None, "Wojtek F", "Wojtek F",
     "W realizacji", "Projektowanie", "Wysoki", "Zolty",
     None, dt.date(2026, 8, 31), None, None, None, None, "PLN", None, None,
     None, None, None, None,
     None, "Projekty: projekt budowlany, projekty na zgloszenie, projekty etapowe, projekty wyburzen.",
     None, U, "[Zrodlo: Wojtek F] Zespol: Wojtek F (lider), Marcin, Gosia + ktos do pomocy (Gosia ma 3 tyg. urlopu). Bardzo obszerny projekt, bardzo krotki czas realizacji - rozne etapy najlepiej przed 20.07.2026, calosc maks. do 31.08.2026."],

    ["GK-01", "99_OPS - Bemowo", "Projekt wykonawczy", "Nadzor autorski", None, "Grzegorz K", "Grzegorz K",
     "W realizacji", "Budowa", None, "Zielony",
     None, dt.date(2026, 12, 31), None, None, None, None, "PLN", None, None,
     "Bemowo", "Warszawa", None, None,
     None, "Nadzor autorski w fazie budowy (projekt wykonawczy).",
     None, U, "[Zrodlo: Grzegorz K] Korespondencja i rozmowy telefoniczne z budowa, odpowiedzi na zapytania o informacje projektowa, wymagane zmiany w dokumentacji, uzupelnianie brakow, konsultowanie i akceptowanie rozwiazan zamiennych, kontakt z branzami, sporadyczne wizyty na budowie i koordynacjach. Termin zakonczenia prac przez GW: do konca 2026."],

    ["GK-02", "109_UCSL - Lublin", "Projekt wykonawczy", "Nadzor autorski", None, "Grzegorz K", "Grzegorz K",
     "W realizacji", "Budowa", None, "Zielony",
     None, dt.date(2026, 12, 31), None, None, None, None, "PLN", None, None,
     None, "Lublin", None, None,
     None, "Nadzor autorski w fazie budowy (projekt wykonawczy).",
     None, U, "[Zrodlo: Grzegorz K] Korespondencja i rozmowy z budowa, zmiany w dokumentacji, sporadyczne wizyty na budowie. Termin zakonczenia prac przez GW: do konca 2026."],

    ["GK-03", "93_SZDF - Siedlce", "Projekt wykonawczy", "Nadzor autorski", None, "Grzegorz K", "Grzegorz K",
     "W realizacji", "Budowa", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Siedlce", None, None,
     None, "Nadzor autorski w fazie budowy (projekt wykonawczy).",
     None, U, "[Zrodlo: Grzegorz K] Korespondencja i rozmowy telefoniczne z inwestorem (nie budowa). Termin zakonczenia prac przez Inwestora nieznany."],

    ["GK-04", "90_BGM - Grodzisk Mazowiecki", "Projekt wykonawczy", "Nadzor autorski", None, "Grzegorz K", "Grzegorz K",
     "W realizacji", "Pozwolenia/Przetarg", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Grodzisk Mazowiecki", None, None,
     None, "Po pozwoleniu na budowe, przed przetargiem na Generalnego Wykonawce.",
     None, U, "[Zrodlo: Grzegorz K] Zespol: Grzegorz K (samodzielnie), wspomagajaco Malgorzata LB. Kontakt z inwestorem, ewentualne poprawki i uzupelnienia w przekazanej dokumentacji (projekt techniczny, zestawienie wyposazenia), odpowiedzi na pytania uczestnikow przetargu. Termin ogloszenia przetargu - wkrotce."],

    ["GK-05", "91_BSO - Sochaczew", "Projekt wykonawczy", "Nadzor autorski", None, "Grzegorz K", "Grzegorz K",
     "W realizacji", "Pozwolenia/Przetarg", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Sochaczew", None, None,
     None, "Po pozwoleniu na budowe, przed przetargiem na Generalnego Wykonawce.",
     None, U, "[Zrodlo: Grzegorz K] Zespol: Grzegorz K (samodzielnie), wspomagajaco Malgorzata LB. Jak wyzej (kontakt z inwestorem, poprawki dokumentacji, odpowiedzi na pytania oferentow). Termin ogloszenia przetargu - wkrotce."],

    ["GK-06", "123_PRB - Bemowo", "Konkurs", "Uczestnik konkursu", None, None, None,
     "W realizacji", "Konkurs - etap studialny", "Wysoki", "Zolty",
     None, dt.date(2026, 10, 9), None, None, None, None, "PLN", None, None,
     "Bemowo", "Warszawa", None, None,
     None, "Konkurs architektoniczny, etap studialny zlozony.",
     None, U, "[Zrodlo: Grzegorz K] Zespol: Toan C, Weronika S (bez formalnego kierownika w notatce). Organizacja pracy zespolu, podzial i przydzial zadan, koordynacja prac, konsultacje i analizy rozwiazan, przygotowanie programu budynku, ukladu funkcjonalnego i komunikacyjnego, sklad calosci projektu, opis, dokumenty. Wyniki najpozniej do 17.07.2026 (moga byc wczesniej ze wzgledu na mala liczbe zlozonych projektow). Jesli II etap: harmonogram prac i nadzor nad prowadzeniem, zlozenie prac II etapu do 9.10.2026."],

    ["MC-01", "LKW_LAKOWA - Siedlce", "Projekt techniczny (PT)", "Koordynacja branzowa", None, "Monika Ch", "Monika Ch",
     "W realizacji", "Zakonczenie", None, "Zielony",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Siedlce", None, None,
     None, "Projekt techniczny (PT) po oddaniu - poprawki i koordynacja z branza instalacji sanitarnych.",
     None, U, "[Zrodlo: Monika Ch] Zespol: Monika Ch."],

    ["MC-02", "WMA_MARZYCIELI - Warszawa", "Projekt budowlany", "Projektant wiodacy", None, "Monika Ch", "Monika Ch",
     "W realizacji", "Pozwolenia/Przetarg", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Warszawa", None, None,
     None, "Projekt budowlany (PB) - uzupelnienia do uwag urzedu; zbieranie materialow do odpowiedzi na uwagi do PB.",
     None, U, "[Zrodlo: Monika Ch] Zespol: Monika Ch (lider), Mikolaj Z (rozmowy z inwestorem), Beata Zdral (dochodzaco)."],

    ["MC-03", "RMS_REMISZEWSKA / MYSZKOWSKA - Warszawa", "Projekt budowlany", "Projektant wiodacy", None, "Monika Ch", "Monika Ch",
     "W realizacji", "Zakonczenie", None, "Zielony",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Warszawa", None, None,
     None, "Projekt budowlany (PB) - zakonczenie fazy projektowej (zmiany do PB).",
     None, U, "[Zrodlo: Monika Ch] Zespol: Monika Ch (lider), Beata Zdral (dochodzaco). Pozostaje zatwierdzenie przez inwestora i wydruk."],

    ["MC-04", "GRA_GRABIANOW - Siedlce", "Projekt koncepcyjny", "Projektant wiodacy", None, "Monika Ch", "Monika Ch",
     "Wstrzymany", "Koncepcja", None, "Zolty",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Siedlce", None, None,
     None, "Projekt koncepcyjny (PK) przeslany do inwestora.",
     None, U, "[Zrodlo: Monika Ch] Zespol: Monika Ch. Czekamy na decyzje inwestora o dalszych krokach w projekcie."],

    ["MC-05", "TBS Pomorska - Warszawa", "Projekt koncepcyjny", "Projektant wiodacy", "Mieszkaniowy", "Monika Ch", "Monika Ch",
     "Planowanie", "Koncepcja", None, "Zielony",
     None, None, None, None, None, None, "PLN", None, None,
     None, "Warszawa", None, None,
     "TBS (nazwa do potwierdzenia)", "Projekt koncepcyjny (PK) - zapoznanie sie z projektem.",
     None, U, "[Zrodlo: Monika Ch] Zespol: Monika Ch + zespol do skompletowania. Duzy temat."],
]
# Wstrzykniecie wartosci Tagi/Przychod_* (None) w istniejace wiersze - naglowki juz sa
# poprawnie zdefiniowane w proj_headers powyzej (przed write_header). Puste, bo tagi/przychody
# nie byly czescia notatek zrodlowych - do uzupelnienia w dashboardzie.
_TAGI_IDX = proj_headers.index("Tagi")
for _row in proj_rows:
    _row.insert(_TAGI_IDX, None)

_PRZYCHOD_IDX = proj_headers.index("Przychod_planowany")
for _row in proj_rows:
    _row.insert(_PRZYCHOD_IDX, None)
    _row.insert(_PRZYCHOD_IDX + 1, None)

assert all(len(_row) == len(proj_headers) for _row in proj_rows), "Projekty: niezgodna liczba kolumn po wstrzyknieciu Tagi/Przychod"

last_proj = write_rows(ws_proj, proj_rows)
add_table(ws_proj, "TabProjekty", proj_headers, 1, last_proj)
autosize(ws_proj, [10, 32, 20, 18, 14, 14, 14, 14, 20, 10, 10, 30, 14, 16, 16, 12, 14, 14, 8, 16, 16, 16, 16, 22, 14, 14, 14, 22, 46, 16, 16, 60])

for col_letter, key in [("C", "Typy_projektow"), ("D", "Funkcje_biura"), ("E", "Segmenty"), ("H", "Statusy_projektu"),
                         ("I", "Fazy"), ("J", "Priorytety"), ("K", "RAG")]:
    dropdown(ws_proj, col_letter, 2, 500, dict_range(key))
dropdown(ws_proj, "F", 2, 500, f"$B$2:$B${len(team_rows)+1}", sheet="Zespol")
dropdown(ws_proj, "G", 2, 500, f"$B$2:$B${len(team_rows)+1}", sheet="Zespol")

for r in range(2, last_proj + 1):
    ws_proj.cell(row=r, column=proj_headers.index("Procent_postepu") + 1).number_format = "0%"
    for col in ("Budzet_calkowity", "Budzet_wydany", "Przychod_planowany", "Przychod_rzeczywisty", "Szacowane_roboczogodziny"):
        ws_proj.cell(row=r, column=proj_headers.index(col) + 1).number_format = "#,##0"
    ws_proj.cell(row=r, column=proj_headers.index("Stawka_godzinowa_srednia") + 1).number_format = "#,##0.00"
    for col in ("Data_rozpoczecia", "Data_zakonczenia_planowana", "Data_zakonczenia_rzeczywista", "Data_ostatniej_aktualizacji"):
        ws_proj.cell(row=r, column=proj_headers.index(col) + 1).number_format = "yyyy-mm-dd"

rag_col = get_column_letter(proj_headers.index("RAG_Status") + 1)
rng = f"{rag_col}2:{rag_col}{last_proj}"
ws_proj.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Zielony"'], fill=FILL_GREEN, font=FONT_GREEN))
ws_proj.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Zolty"'], fill=FILL_YELLOW, font=FONT_YELLOW))
ws_proj.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Czerwony"'], fill=FILL_RED, font=FONT_RED))

status_col = get_column_letter(proj_headers.index("Status") + 1)
rng2 = f"{status_col}2:{status_col}{last_proj}"
ws_proj.conditional_formatting.add(rng2, CellIsRule(operator="equal", formula=['"Zakonczony"'], fill=FILL_GREEN, font=FONT_GREEN))
ws_proj.conditional_formatting.add(rng2, CellIsRule(operator="equal", formula=['"Wstrzymany"'], fill=FILL_RED, font=FONT_RED))
ws_proj.conditional_formatting.add(rng2, CellIsRule(operator="equal", formula=['"Anulowany"'], fill=FILL_RED, font=FONT_RED))

ws_proj.sheet_properties.tabColor = "1F3864"

# ===========================================================================
# 4) PRZYPISANIA (n:n Projekty <-> Zespol)
# ===========================================================================
ws_assign = add_sheet("Przypisania")
assign_headers = ["ID_Przypisania", "ID_Projektu", "ID_Osoby", "Rola_w_projekcie",
                   "Procent_zaangazowania", "Data_od", "Data_do", "Status"]
write_header(ws_assign, assign_headers)


def A(pid, oid, rola, status="Aktywny"):
    return [None, pid, oid, rola, None, None, None, status]


assign_rows_raw = [
    A("JB-01", "P01", "Kierownik projektu"), A("JB-01", "P02", "Czlonek zespolu"), A("JB-01", "P03", "Czlonek zespolu"),
    A("JB-02", "P01", "Kierownik projektu"), A("JB-02", "P02", "Czlonek zespolu"),
    A("JB-03", "P01", "Kierownik projektu"),
    A("JB-05", "P01", "Kierownik projektu"), A("JB-05", "P02", "Czlonek zespolu"),
    A("WF-01", "P04", "Kierownik projektu"),
    A("WF-02", "P04", "Kierownik projektu"), A("WF-02", "P05", "Wsparcie/Konsultant"),
    A("WF-03", "P04", "Kierownik projektu"), A("WF-03", "P05", "Czlonek zespolu"),
    A("WF-04", "P04", "Kierownik projektu"), A("WF-04", "P06", "Czlonek zespolu"), A("WF-04", "P05", "Wsparcie/Konsultant"),
    A("WF-05", "P04", "Kierownik projektu"), A("WF-05", "P05", "Czlonek zespolu"), A("WF-05", "P07", "Czlonek zespolu"),
    A("GK-01", "P08", "Kierownik projektu"),
    A("GK-02", "P08", "Kierownik projektu"),
    A("GK-03", "P08", "Kierownik projektu"),
    A("GK-04", "P08", "Kierownik projektu"), A("GK-04", "P09", "Wsparcie/Konsultant"),
    A("GK-05", "P08", "Kierownik projektu"), A("GK-05", "P09", "Wsparcie/Konsultant"),
    A("GK-06", "P02", "Czlonek zespolu"), A("GK-06", "P03", "Czlonek zespolu"),
    A("MC-01", "P10", "Kierownik projektu"),
    A("MC-02", "P10", "Kierownik projektu"), A("MC-02", "P11", "Wsparcie/Konsultant"), A("MC-02", "P12", "Wsparcie/Konsultant"),
    A("MC-03", "P10", "Kierownik projektu"), A("MC-03", "P12", "Wsparcie/Konsultant"),
    A("MC-04", "P10", "Kierownik projektu"),
    A("MC-05", "P10", "Kierownik projektu"),
]
for i, row in enumerate(assign_rows_raw, start=1):
    row[0] = f"ASG{i:03d}"
last_assign = write_rows(ws_assign, assign_rows_raw)
add_table(ws_assign, "TabPrzypisania", assign_headers, 1, last_assign)
autosize(ws_assign, [14, 12, 10, 20, 12, 14, 14, 12])
dropdown(ws_assign, "B", 2, 500, f"$A$2:$A${len(proj_rows)+1}", sheet="Projekty")
dropdown(ws_assign, "C", 2, 500, f"$A$2:$A${len(team_rows)+1}", sheet="Zespol")
dropdown(ws_assign, "D", 2, 500, dict_range("Role_w_projekcie"))
dropdown(ws_assign, "H", 2, 500, "={\"Aktywny\",\"Zakonczony\"}".replace("=", ""))
for r in range(2, last_assign + 1):
    ws_assign.cell(row=r, column=5).number_format = '0"%"'
ws_assign.sheet_properties.tabColor = "1F3864"

# ===========================================================================
# 5) HARMONOGRAM (dane pod Gantta) - CELOWO PUSTY: struktura + walidacja, bez danych
# ===========================================================================
ws_sched = add_sheet("Harmonogram")
sched_headers = [
    "ID_Zadania", "ID_Projektu", "Nazwa_zadania", "Kategoria", "ID_Osoby_odpowiedzialnej",
    "Data_start_plan", "Data_koniec_plan", "Data_start_rzeczywista", "Data_koniec_rzeczywista",
    "Procent_ukonczenia", "ID_Zadania_poprzedzajacego", "Kamien_milowy", "Status", "Priorytet", "Uwagi",
]
write_header(ws_sched, sched_headers)
sched_rows_raw = []
last_sched = safe_last(write_rows(ws_sched, sched_rows_raw))
add_table(ws_sched, "TabHarmonogram", sched_headers, 1, last_sched if sched_rows_raw else 0)
autosize(ws_sched, [10, 12, 30, 16, 14, 14, 14, 16, 16, 12, 14, 12, 14, 10, 40])
dropdown(ws_sched, "B", 2, 500, f"$A$2:$A${len(proj_rows)+1}", sheet="Projekty")
dropdown(ws_sched, "D", 2, 500, dict_range("Kategorie_zadan"))
dropdown(ws_sched, "E", 2, 500, f"$A$2:$A${len(team_rows)+1}", sheet="Zespol")
dropdown(ws_sched, "L", 2, 500, "={\"Tak\",\"Nie\"}".replace("=", ""))
dropdown(ws_sched, "M", 2, 500, dict_range("Status_zadania"))
dropdown(ws_sched, "N", 2, 500, dict_range("Priorytety"))
for r in range(2, last_sched + 1):
    ws_sched.cell(row=r, column=10).number_format = "0%"
    for col in (6, 7, 8, 9):
        ws_sched.cell(row=r, column=col).number_format = "yyyy-mm-dd"

status_col_s = get_column_letter(13)
rngS = f"{status_col_s}2:{status_col_s}{last_sched}"
ws_sched.conditional_formatting.add(rngS, CellIsRule(operator="equal", formula=['"Zakonczone"'], fill=FILL_GREEN, font=FONT_GREEN))
ws_sched.conditional_formatting.add(rngS, CellIsRule(operator="equal", formula=['"Opoznione"'], fill=FILL_RED, font=FONT_RED))
ws_sched.conditional_formatting.add(rngS, CellIsRule(operator="equal", formula=['"W trakcie"'], fill=FILL_YELLOW, font=FONT_YELLOW))

GANTT_START = dt.date(2024, 1, 1)
N_MONTHS = 48
gantt_first_col = len(sched_headers) + 2
month_header_row = 1
for m in range(N_MONTHS):
    year = GANTT_START.year + (GANTT_START.month - 1 + m) // 12
    month = (GANTT_START.month - 1 + m) % 12 + 1
    col = gantt_first_col + m
    cell = ws_sched.cell(row=month_header_row, column=col, value=dt.date(year, month, 1))
    cell.number_format = "mmm-yy"
    cell.font = Font(size=8, bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", textRotation=90)
    ws_sched.column_dimensions[get_column_letter(col)].width = 3.4

gantt_last_col = gantt_first_col + N_MONTHS - 1
start_col_letter = get_column_letter(sched_headers.index("Data_start_plan") + 1)
end_col_letter = get_column_letter(sched_headers.index("Data_koniec_plan") + 1)
status_letter = get_column_letter(sched_headers.index("Status") + 1)
first_g = get_column_letter(gantt_first_col)
last_g = get_column_letter(gantt_last_col)
gantt_range = f"{first_g}2:{last_g}{last_sched}"


def gantt_rule(status_value, fill):
    formula = (f'=AND(${start_col_letter}2<=EOMONTH({first_g}$1,0),${end_col_letter}2>={first_g}$1,'
               f'${status_letter}2="{status_value}")')
    return FormulaRule(formula=[formula], fill=fill)


ws_sched.conditional_formatting.add(gantt_range, gantt_rule("Zakonczone", FILL_GREEN))
ws_sched.conditional_formatting.add(gantt_range, gantt_rule("W trakcie", PatternFill("solid", fgColor="9DC3E6")))
ws_sched.conditional_formatting.add(gantt_range, gantt_rule("Opoznione", FILL_RED))
ws_sched.conditional_formatting.add(gantt_range, gantt_rule("Nie rozpoczete", PatternFill("solid", fgColor="D9D9D9")))
ws_sched.sheet_properties.tabColor = "BF8F00"
ws_sched.cell(row=1, column=1).comment = Comment(
    "Arkusz celowo pusty (brak jeszcze danych o etapach/zadaniach) - struktura, walidacja i "
    "warunkowy Gantt sa gotowe do uzupelnienia na biezaco przez zespol (recznie lub przez dashboard).", "PMO")

# ===========================================================================
# 5b) ZADANIA (TICKETY) - CELOWO PUSTY (granularne zadania per osoba, do zbudowania w dashboardzie)
# ===========================================================================
ws_tk = add_sheet("Zadania_Tickety")
tk_headers = ["ID_Tickietu", "ID_Projektu", "ID_Etapu", "Tytul", "Opis", "ID_Osoby_przypisanej",
              "ID_Podwykonawcy", "Wycena_podwykonawcy",
              "Data_utworzenia", "Termin", "Szacowane_roboczogodziny", "Rzeczywiste_roboczogodziny",
              "Priorytet", "Status", "Data_zakonczenia"]
write_header(ws_tk, tk_headers)
tk_rows_raw = []
last_tk = safe_last(write_rows(ws_tk, tk_rows_raw))
autosize(ws_tk, [12, 12, 12, 40, 46, 16, 16, 16, 14, 14, 14, 16, 12, 14, 16])
dropdown(ws_tk, "B", 2, 500, f"$A$2:$A${len(proj_rows)+1}", sheet="Projekty")
dropdown(ws_tk, "F", 2, 500, f"$A$2:$A${len(team_rows)+1}", sheet="Zespol")
dropdown(ws_tk, "G", 2, 500, "$A$2:$A$300", sheet="Podwykonawcy")
dropdown(ws_tk, "M", 2, 500, dict_range("Priorytety"))
dropdown(ws_tk, "N", 2, 500, dict_range("Statusy_tickietow"))
for r in range(2, last_tk + 1):
    ws_tk.cell(row=r, column=8).number_format = "#,##0.00"
    ws_tk.cell(row=r, column=9).number_format = "yyyy-mm-dd"
    ws_tk.cell(row=r, column=10).number_format = "yyyy-mm-dd"
    ws_tk.cell(row=r, column=11).number_format = "0"
    ws_tk.cell(row=r, column=12).number_format = "0"
    ws_tk.cell(row=r, column=15).number_format = "yyyy-mm-dd"
status_col_tk = get_column_letter(14)
rngTk = f"{status_col_tk}2:{status_col_tk}{last_tk}"
ws_tk.conditional_formatting.add(rngTk, CellIsRule(operator="equal", formula=['"Zrobione"'], fill=FILL_GREEN, font=FONT_GREEN))
ws_tk.conditional_formatting.add(rngTk, CellIsRule(operator="equal", formula=['"Zablokowane"'], fill=FILL_RED, font=FONT_RED))
ws_tk.conditional_formatting.add(rngTk, CellIsRule(operator="equal", formula=['"W trakcie"'], fill=FILL_YELLOW, font=FONT_YELLOW))
ws_tk.sheet_properties.tabColor = "BF8F00"
ws_tk.cell(row=1, column=1).comment = Comment(
    "Granularne zadania (tickety) przypisane do konkretnej osoby (zespol wewnetrzny) LUB podwykonawcy, "
    "z terminem i roboczogodzinami. Rozni sie od 'Harmonogram' (etapy/fazy projektu pod Gantta) - ticket "
    "moze (opcjonalnie, ID_Etapu) byc powiazany z etapem, ale to osobny, drobniejszy poziom pracy. Gdy "
    "ustawiony jest ID_Podwykonawcy, Wycena_podwykonawcy wchodzi do kosztu projektu i marzowosci "
    "(liczone w dashboardzie) - niezaleznie od Rzeczywiste_roboczogodziny x Stawka_godzinowa dla "
    "zadan zespolu wewnetrznego (arkusz Zespol). Arkusz celowo pusty - do uzupelniania na biezaco "
    "przez zespol w dashboardzie.", "PMO")

# ===========================================================================
# 6) KAMIENIE MILOWE - CELOWO PUSTY
# ===========================================================================
ws_mile = add_sheet("Kamienie_milowe")
mile_headers = ["ID_Kamienia", "ID_Projektu", "Nazwa_kamienia", "Data_planowana", "Data_rzeczywista",
                "Status", "ID_Osoby_odpowiedzialnej"]
write_header(ws_mile, mile_headers)
mile_rows_raw = []
last_mile = safe_last(write_rows(ws_mile, mile_rows_raw))
autosize(ws_mile, [12, 12, 34, 16, 16, 14, 14])
dropdown(ws_mile, "B", 2, 500, f"$A$2:$A${len(proj_rows)+1}", sheet="Projekty")
dropdown(ws_mile, "F", 2, 500, "={\"Planowane\",\"W trakcie\",\"Zakonczone\",\"Zagrozone\"}".replace("=", ""))
dropdown(ws_mile, "G", 2, 500, f"$A$2:$A${len(team_rows)+1}", sheet="Zespol")
for r in range(2, last_mile + 1):
    ws_mile.cell(row=r, column=4).number_format = "yyyy-mm-dd"
    ws_mile.cell(row=r, column=5).number_format = "yyyy-mm-dd"
ws_mile.sheet_properties.tabColor = "1F3864"

# ===========================================================================
# 6b) PODWYKONAWCY - CELOWO PUSTY (biblioteka do zbudowania)
# ===========================================================================
ws_sub = add_sheet("Podwykonawcy")
sub_headers = ["ID_Podwykonawcy", "Nazwa", "Branza", "Typ_wspolpracy", "Osoba_kontaktowa",
               "Email", "Telefon", "NIP", "Miasto", "Ocena", "Status", "Uwagi"]
write_header(ws_sub, sub_headers)
sub_rows_raw = []
last_sub = safe_last(write_rows(ws_sub, sub_rows_raw))
autosize(ws_sub, [16, 32, 24, 20, 20, 30, 14, 14, 14, 12, 14, 40])
dropdown(ws_sub, "C", 2, 300, dict_range("Branze_podwykonawcow"))
dropdown(ws_sub, "D", 2, 300, dict_range("Typy_wspolpracy_podwykonawcow"))
dropdown(ws_sub, "J", 2, 300, dict_range("Oceny_podwykonawcow"))
dropdown(ws_sub, "K", 2, 300, dict_range("Statusy_podwykonawcow"))
ws_sub.sheet_properties.tabColor = "BF8F00"
ws_sub.cell(row=1, column=1).comment = Comment(
    "Biblioteka podwykonawcow/branzystow - niezalezna od projektow. Pusta na start, "
    "do uzupelniania na biezaco (recznie lub przez dashboard).", "PMO")

# ===========================================================================
# 6c) PRZYPISANIA PODWYKONAWCOW DO PROJEKTOW - CELOWO PUSTY
# ===========================================================================
ws_suba = add_sheet("Przypisania_Podwykonawcow")
suba_headers = ["ID_Przypisania_Podw", "ID_Projektu", "ID_Podwykonawcy", "Branza", "Zakres_prac",
                 "Data_od", "Data_do", "Wartosc_umowy", "Waluta", "Status", "Uwagi"]
write_header(ws_suba, suba_headers)
suba_rows_raw = []
last_suba = safe_last(write_rows(ws_suba, suba_rows_raw))
autosize(ws_suba, [16, 12, 16, 24, 44, 14, 14, 14, 8, 14, 30])
dropdown(ws_suba, "B", 2, 500, f"$A$2:$A${len(proj_rows)+1}", sheet="Projekty")
dropdown(ws_suba, "C", 2, 500, f"$A$2:$A${safe_last(len(sub_rows_raw)+1)}", sheet="Podwykonawcy")
dropdown(ws_suba, "D", 2, 500, dict_range("Branze_podwykonawcow"))
dropdown(ws_suba, "J", 2, 500, dict_range("Statusy_przypisania_podwykonawcy"))
for r in range(2, last_suba + 1):
    ws_suba.cell(row=r, column=6).number_format = "yyyy-mm-dd"
    ws_suba.cell(row=r, column=7).number_format = "yyyy-mm-dd"
    ws_suba.cell(row=r, column=8).number_format = "#,##0"
ws_suba.sheet_properties.tabColor = "BF8F00"

# ===========================================================================
# 7) RYZYKA I PROBLEMY - CELOWO PUSTY
# ===========================================================================
ws_risk = add_sheet("Ryzyka_i_Problemy")
risk_headers = ["ID", "ID_Projektu", "Typ", "Opis", "Kategoria", "Prawdopodobienstwo", "Wplyw",
                 "Priorytet", "ID_Osoby_wlasciciela", "Plan_mitygacji", "Status",
                 "Data_identyfikacji", "Data_zamkniecia"]
write_header(ws_risk, risk_headers)
risk_rows_raw = []
last_risk = safe_last(write_rows(ws_risk, risk_rows_raw))
autosize(ws_risk, [10, 12, 44, 16, 14, 10, 10, 14, 40, 12, 14, 14])
dropdown(ws_risk, "B", 2, 500, f"$A$2:$A${len(proj_rows)+1}", sheet="Projekty")
dropdown(ws_risk, "C", 2, 500, dict_range("Typ_ryzyka"))
dropdown(ws_risk, "E", 2, 500, dict_range("Kategorie_ryzyk"))
dropdown(ws_risk, "F", 2, 500, dict_range("Priorytety"))
dropdown(ws_risk, "G", 2, 500, dict_range("Priorytety"))
dropdown(ws_risk, "H", 2, 500, dict_range("Priorytety"))
dropdown(ws_risk, "I", 2, 500, f"$A$2:$A${len(team_rows)+1}", sheet="Zespol")
dropdown(ws_risk, "K", 2, 500, dict_range("Status_ryzyka"))
for r in range(2, last_risk + 1):
    ws_risk.cell(row=r, column=12).number_format = "yyyy-mm-dd"
    ws_risk.cell(row=r, column=13).number_format = "yyyy-mm-dd"
ws_risk.sheet_properties.tabColor = "C00000"

# ===========================================================================
# 8) RAPORTY STATUSOWE (historia - trend) - CELOWO PUSTY
# ===========================================================================
ws_stat = add_sheet("Raporty_statusowe")
stat_headers = ["Data_raportu", "ID_Projektu", "RAG_Status", "Procent_postepu",
                "Budzet_wydany_skumulowany", "Kluczowe_osiagniecia", "Kluczowe_problemy",
                "Nastepne_kroki", "Autor_raportu"]
write_header(ws_stat, stat_headers)
stat_rows_raw = []
last_stat = safe_last(write_rows(ws_stat, stat_rows_raw))
autosize(ws_stat, [14, 12, 12, 14, 20, 40, 40, 40, 18])
dropdown(ws_stat, "B", 2, 500, f"$A$2:$A${len(proj_rows)+1}", sheet="Projekty")
dropdown(ws_stat, "C", 2, 500, dict_range("RAG"))
for r in range(2, last_stat + 1):
    ws_stat.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    ws_stat.cell(row=r, column=4).number_format = "0%"
    ws_stat.cell(row=r, column=5).number_format = "#,##0"
ws_stat.sheet_properties.tabColor = "808080"

# ===========================================================================
# 9) KARTA PROJEKTU (wzor do wydruku / rozbudowy recznej)
# ===========================================================================
ws_card = add_sheet("Karta_Projektu")
ws_card.sheet_view.showGridLines = False
title = ws_card.cell(row=1, column=1, value="KARTA PROJEKTU - WZOR")
title.font = Font(size=16, bold=True, color="1F3864")
ws_card.merge_cells("A1:D1")
note = ws_card.cell(row=2, column=1,
                     value="Wzor jednostronicowej karty do wydruku/omowienia na komitecie sterujacym. "
                           "Zywa, interaktywna karta projektu (z automatycznym wyborem projektu) jest w dashboardzie HTML: "
                           "dashboard/index.html -> zakladka 'Projekty' -> klik w kafelek projektu.")
note.font = Font(italic=True, size=9, color="808080")
ws_card.merge_cells("A2:F2")
ws_card.row_dimensions[2].height = 30
note.alignment = Alignment(wrap_text=True, vertical="top")

fields = [
    "Nazwa projektu", "ID projektu", "Typ projektu", "Funkcja biura", "Segment", "Tagi", "Lokalizacja / Miasto",
    "Owner", "Kierownik projektu", "Status", "Faza", "Priorytet", "RAG Status",
    "Data rozpoczecia", "Data zakonczenia (plan)", "Procent postepu",
    "Budzet calkowity", "Budzet wydany", "Przychod planowany / rzeczywisty", "Marza / Mark-up",
    "Inwestor / Klient",
    "Opis / Zakres", "Kluczowe kamienie milowe", "Kluczowe ryzyka", "Zespol projektowy",
    "Komentarz PMO",
]
r = 4
for f in fields:
    lbl = ws_card.cell(row=r, column=1, value=f)
    lbl.font = Font(bold=True, size=10)
    lbl.fill = PatternFill("solid", fgColor="EDEDED")
    lbl.alignment = Alignment(vertical="top")
    val = ws_card.cell(row=r, column=2)
    val.border = BORDER
    val.alignment = Alignment(wrap_text=True, vertical="top")
    ws_card.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws_card.row_dimensions[r].height = 18
    r += 1
autosize(ws_card, [24, 16, 16, 16, 16, 16])
ws_card.sheet_properties.tabColor = "808080"

# ===========================================================================
# 10) PULPIT (KPI portfela - formuly, dziala od razu po otwarciu w Excelu)
# ===========================================================================
ws_dash = add_sheet("Pulpit")
ws_dash.sheet_view.showGridLines = False
t = ws_dash.cell(row=1, column=1, value="PULPIT PORTFELA PROJEKTOW")
t.font = Font(size=16, bold=True, color="1F3864")
ws_dash.merge_cells("A1:D1")
sub = ws_dash.cell(row=2, column=1, value="Formuly licza sie automatycznie w Excelu na podstawie arkusza 'Projekty'. "
                                           "Pelny interaktywny dashboard: dashboard/index.html.")
sub.font = Font(italic=True, size=9, color="808080")
ws_dash.merge_cells("A2:F2")

kpi_start = 4
kpis = [
    ("Liczba projektow ogolem", f"=COUNTA(Projekty!A2:A{last_proj})"),
    ("W realizacji", f'=COUNTIF(Projekty!G2:G{last_proj},"W realizacji")'),
    ("Planowanie", f'=COUNTIF(Projekty!G2:G{last_proj},"Planowanie")'),
    ("Wstrzymane", f'=COUNTIF(Projekty!G2:G{last_proj},"Wstrzymany")'),
    ("Zakonczone", f'=COUNTIF(Projekty!G2:G{last_proj},"Zakonczony")'),
    ("Projekty z czerwonym RAG", f'=COUNTIF(Projekty!J2:J{last_proj},"Czerwony")'),
    ("Projekty z zoltym RAG", f'=COUNTIF(Projekty!J2:J{last_proj},"Zolty")'),
    ("Budzet calkowity (suma)", f"=SUM(Projekty!O2:O{last_proj})"),
    ("Budzet wydany (suma)", f"=SUM(Projekty!P2:P{last_proj})"),
    ("Suma szacowanych roboczogodzin", f"=SUM(Projekty!R2:R{last_proj})"),
    ("Otwarte ryzyka/problemy", f'=COUNTIF(Ryzyka_i_Problemy!K2:K{last_risk},"Otwarte")'),
    ("Nadchodzace kamienie (90 dni)", f"=COUNTIFS(Kamienie_milowe!D2:D{last_mile},\">=\"&TODAY(),Kamienie_milowe!D2:D{last_mile},\"<=\"&TODAY()+90)"),
    ("Podwykonawcy w bibliotece", f"=COUNTA(Podwykonawcy!A2:A{last_sub})"),
    ("Aktywne przypisania podwykonawcow", f'=COUNTIF(Przypisania_Podwykonawcow!J2:J{last_suba},"Aktywny")'),
    ("Tickety opoznione", f'=COUNTIFS(Zadania_Tickety!J2:J{last_tk},"<"&TODAY(),Zadania_Tickety!N2:N{last_tk},"<>Zrobione")-COUNTIFS(Zadania_Tickety!J2:J{last_tk},"<"&TODAY(),Zadania_Tickety!N2:N{last_tk},"Zarchiwizowane")'),
    ("Tickety otwarte (poza Zrobione/Zarchiwizowane)", f'=COUNTA(Zadania_Tickety!A2:A{last_tk})-COUNTIF(Zadania_Tickety!N2:N{last_tk},"Zrobione")-COUNTIF(Zadania_Tickety!N2:N{last_tk},"Zarchiwizowane")'),
]
r = kpi_start
for label, formula in kpis:
    lbl = ws_dash.cell(row=r, column=1, value=label)
    lbl.font = Font(bold=True, size=10)
    val = ws_dash.cell(row=r, column=3, value=formula)
    val.font = Font(size=12, bold=True, color="1F3864")
    if "Budzet" in label or "roboczogodzin" in label.lower():
        val.number_format = "#,##0"
    r += 1
autosize(ws_dash, [32, 4, 16])

desired_order = ["Pulpit", "Projekty", "Karta_Projektu", "Zespol", "Przypisania", "Harmonogram",
                  "Zadania_Tickety", "Kamienie_milowe", "Podwykonawcy", "Przypisania_Podwykonawcow",
                  "Ryzyka_i_Problemy", "Raporty_statusowe", "Slowniki"]
wb._sheets = [wb[name] for name in desired_order]
wb.active = 0
out_path = "Baza_Projektow.xlsx"
wb.save(out_path)
print(f"Zapisano: {out_path}")
print("Sheety:", wb.sheetnames)
